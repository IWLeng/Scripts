import pandas as pd
import requests
import time
import os
import re
from openpyxl import load_workbook
from tqdm import tqdm

# --- Configuration ---
API_KEY = "AIzaSyD_5Aye31mk3p_N3ZA3JXVdzHLP16QElHA"  # Replace with your key
REQUEST_TIMEOUT = 30  # API timeout in seconds

# --- Translation Functions ---
def translate_batch(texts, source_language, target_language):
    """Translate a batch of texts using Google Cloud Translation API."""
    url = "https://translation.googleapis.com/language/translate/v2"
    
    # Prepare the request data as JSON in the body
    data = {
        "q": texts,
        "source": source_language,
        "target": target_language,
    }
    
    params = {
        "key": API_KEY,
    }
    
    headers = {
        "Content-Type": "application/json",
    }
    
    try:
        response = requests.post(
            url,
            params=params,
            json=data,
            headers=headers,
            timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()
        return [t["translatedText"] for t in response.json()["data"]["translations"]]
    except requests.exceptions.RequestException as e:
        raise Exception(f"API request failed: {e}")

# --- Helper Functions ---
def is_non_translatable(text):
    """Check if a cell contains only numbers/symbols."""
    return bool(re.match(r'^[\W\d_]+$', str(text)))

def column_letter_to_index(column_letter):
    """Convert Excel column letter (A, B, ..., AA, AB, etc.) to zero-based index."""
    index = 0
    for i, char in enumerate(reversed(column_letter.upper())):
        index += (ord(char) - ord('A') + 1) * (26 ** i)
    return index - 1  # Convert to zero-based

def count_words(text):
    """Simple word counter (split by whitespace)."""
    return len(str(text).split())

# --- Main Translation Function ---
def translate_xlsx(input_file, source_language="es", target_language="en"):
    # Initialize counters
    translated_words = 0
    failed_words = 0
    
    # Set up file paths
    input_dir = os.path.dirname(input_file)
    input_name = os.path.splitext(os.path.basename(input_file))[0]
    output_file = os.path.join(input_dir, f"{input_name}_translated.xlsx")
    log_file = os.path.join(input_dir, f"{input_name}_log.xlsx")
    summary_file = os.path.join(input_dir, f"{input_name}_summary.txt")

    # Load workbook
    wb = load_workbook(input_file)
    ws = wb.active
    log_data = []

    # --- Column Selection ---
    user_input = input("Enter columns to EXCLUDE from translation (e.g., A,C or AB,AC or leave empty for none): ").strip()
    columns_to_exclude = []
    if user_input:
        # Split by comma and remove whitespace
        column_letters = [col.strip() for col in user_input.split(",")]
        columns_to_exclude = [column_letter_to_index(col) for col in column_letters]

    # --- Translation Process ---
    MAX_CHARS_PER_MINUTE = 600000
    MAX_TEXTS_PER_BATCH = 32  # Reduced from 128 to prevent large requests
    MAX_CHARS_PER_BATCH = 50000  # Reduced from 1000000 to prevent large requests
    total_chars = 0
    start_time = time.time()

    # Prepare texts to translate
    texts = []
    cell_refs = []
    for row in ws.iter_rows():
        for col_idx, cell in enumerate(row):
            if col_idx in columns_to_exclude:
                continue
            if isinstance(cell.value, str) and not is_non_translatable(cell.value):
                texts.append(cell.value)
                cell_refs.append(cell)

    # Process in batches
    with tqdm(total=len(texts), desc="Translating") as pbar:
        i = 0
        while i < len(texts):
            # Create batch
            batch = []
            batch_cells = []
            batch_chars = 0
            while i < len(texts) and len(batch) < MAX_TEXTS_PER_BATCH and batch_chars + len(texts[i]) <= MAX_CHARS_PER_BATCH:
                batch.append(texts[i])
                batch_cells.append(cell_refs[i])
                batch_chars += len(texts[i])
                i += 1

            # Rate limiting
            elapsed = time.time() - start_time
            if total_chars + batch_chars > MAX_CHARS_PER_MINUTE:
                delay = max(60 - elapsed, 0)
                print(f"Rate limit: Waiting {delay:.1f}s...")
                time.sleep(delay)
                total_chars = 0
                start_time = time.time()

            # Translate with retries
            translated = None
            for attempt in range(3):
                try:
                    translated = translate_batch(batch, source_language, target_language)
                    break
                except Exception as e:
                    print(f"Attempt {attempt + 1} failed: {e}")
                    time.sleep(2 ** attempt)

            # Update results
            if translated:
                for cell, text in zip(batch_cells, translated):
                    cell.value = text
                    translated_words += count_words(text)
            else:
                for cell, text in zip(batch_cells, batch):
                    log_data.append([cell.coordinate, text, "Failed"])
                    failed_words += count_words(text)

            total_chars += batch_chars
            pbar.update(len(batch))

    # Save results
    wb.save(output_file)
    pd.DataFrame(log_data, columns=["Cell", "Original", "Translated"]).to_excel(log_file, index=False)

    # --- Final Summary ---
    total_words = translated_words + failed_words
    summary_text = f"""Translation Summary:
File: {input_name}
- Successfully translated: {translated_words:,} words
- Failed translations: {failed_words:,} words
- Total processed: {total_words:,} words

"""

    # Print to console
    print("\n" + summary_text)
    
    # Save to text file
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write(summary_text)
    print(f"Summary file saved to: {summary_file}")

# --- Run Script ---
if __name__ == "__main__":
    input_file = input("Enter path to XLSX file: ").strip('"')
    translate_xlsx(input_file)