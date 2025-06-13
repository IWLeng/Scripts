import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Ask user for file paths
original_file = input("Enter the path to the original Excel file: ").strip()
log_file = input("Enter the path to the log file with translations: ").strip()  # User provides log file path 

# Load log file with translations
log_df = pd.read_excel(log_file, dtype=str)  # Ensure strings are read properly

# Load original Excel file
wb = load_workbook(original_file)

# Insert translations
for _, row in log_df.iterrows():
    cell_id, _, translated_text = row  # Extract values from log

    if pd.isna(translated_text) or translated_text.strip() == "":  # Skip empty translations
        continue

    column_letter = ''.join(filter(str.isalpha, cell_id))  # Extract column (e.g., "B")
    row_number = ''.join(filter(str.isdigit, cell_id))  # Extract row number (e.g., "20")

    if not row_number.isdigit():  # Safety check
        continue

    sheet_name = list(wb.sheetnames)[0]  # Assume single sheet; change if needed
    sheet = wb[sheet_name]

    col_index = column_index_from_string(column_letter)  # Convert to numeric index
    sheet.cell(row=int(row_number), column=col_index, value=translated_text)  # Update cell

# Save updated file
updated_file = original_file.replace(".xlsx", "_translated.xlsx")
wb.save(updated_file)

print(f"✅ Translations inserted successfully!")
print(f"📜 Updated file saved as: {updated_file}")
